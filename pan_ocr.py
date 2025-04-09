# pan_card_extractor.py
import cv2
import numpy as np
import pytesseract
from ultralytics import YOLO
import re
from datetime import datetime
from collections import Counter
import os
import glob
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

class PANProcessor:
    def __init__(self):
        self.class_map = {
            0: "dob",
            1: "father_name",
            2: "name",
            3: "pan_number"
        }
        self.date_formats = [
            (r'(\d{1,2})[/\-\.\s](\d{1,2})[/\-\.\s](\d{4})', 'dmy'),
            (r'(\d{4})[/\-\.\s](\d{1,2})[/\-\.\s](\d{1,2})', 'ymd'),
            (r'(\d{1,2})[/\-\.\s]([A-Za-z]{3,})[/\-\.\s](\d{4})', 'dby'),
        ]
        self.month_map = {
            'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
            'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
        }
        self.excel_headers = ["Name", "Father's Name", "PAN Number", "DOB"]
        self.model = YOLO("best.pt")
        pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

    def preprocess(self, image):
        """Enhance text regions with safety checks"""
        if image is None or image.size == 0:
            return None
        try:
            if len(image.shape) == 3:
                gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            else:
                gray = image
            return cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
        except cv2.error:
            return None

    def _validate_date(self, day, month, year):
        """Check if date components form valid date"""
        try:
            if 1900 <= year <= datetime.now().year + 10:
                datetime(year=year, month=month, day=day)
                return True
        except ValueError:
            return False
        return False

    def _process_dob(self, text):
        """Final optimized date parsing with enhanced validation"""
        text = text.replace('O', '0').replace('o', '0').replace('l', '1')
        text = re.sub(r'[^\d/\-\.A-Za-z]', '', text).strip()
        
        # Enhanced date patterns with priority
        patterns = [
            # DD/MM/YYYY with any separator
            (r'(\d{2})[./\-](\d{2})[./\-](\d{4})', 'dmy'),
            # MM/DD/YYYY with any separator
            (r'(\d{2})[./\-](\d{2})[./\-](\d{4})', 'mdy'),
            # Month name format (DD-Mon-YYYY)
            (r'(\d{1,2})[./\- ]([A-Za-z]{3,})[./\- ](\d{4})', 'dby'),
            # Year-first format
            (r'(\d{4})[./\-](\d{2})[./\-](\d{2})', 'ymd'),
        ]

        best_date = None
        max_confidence = 0
        
        for pattern, fmt in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                try:
                    if fmt == 'dmy':
                        day, month, year = map(int, match.groups())
                    elif fmt == 'mdy':
                        month, day, year = map(int, match.groups())
                    elif fmt == 'dby':
                        day = int(match.group(1))
                        month_str = match.group(2)[:3].lower()
                        month = self.month_map.get(month_str)
                        year = int(match.group(3))
                    elif fmt == 'ymd':
                        year, month, day = map(int, match.groups())
                    
                    # Validate date components
                    if not (1 <= day <= 31 and 1 <= month <= 12 and 1900 <= year <= datetime.now().year):
                        continue
                        
                    # Validate actual date existence
                    datetime(year=year, month=month, day=day)
                    confidence = len(text.strip())  # Longer matches are better
                    
                    if confidence > max_confidence:
                        best_date = f"{day:02d}/{month:02d}/{year}"
                        max_confidence = confidence
                        
                except (ValueError, KeyError, TypeError):
                    continue

        return best_date or ""

    def _process_pan(self, text):
        """Robust PAN validation with OCR correction"""
        CHAR_MAP = {'0':'O','1':'I','2':'Z','4':'A','5':'S','8':'B',
                   'B':'8','D':'0','I':'1','O':'0','S':'5','Z':'2'}
        cleaned = re.sub(r'[^A-Z0-9]', '', text.upper())
        if len(cleaned) != 10:
            return ""
        
        parts = [
            [c if c.isalpha() else CHAR_MAP.get(c, c) for c in cleaned[:5]],
            [c if c.isdigit() else CHAR_MAP.get(c, c) for c in cleaned[5:9]],
            [CHAR_MAP.get(cleaned[9], '') if not cleaned[9].isalpha() else cleaned[9]]
        ]
        
        pan = ''.join([''.join(p) for p in parts])
        if re.match(r'^[A-Z]{5}[0-9]{4}[A-Z]$', pan):
            return pan
        return ""

    def _process_name(self, text):
        """Validate names with OCR correction"""
        text = re.sub(r'[^A-Za-z\s\']', '', text)
        if re.fullmatch(r'^[Oo]+$', text.strip()):
            return ""
        text = re.sub(r'\s+', ' ', text).strip().title()
        return text if len(text) >= 2 and not any(c.isdigit() for c in text) else ""
        '''text = re.sub(r'[Oo]', '', text)
        text = re.sub(r'[^A-Za-z\s\'\-]', '', text)
        text = re.sub(r'\s+', ' ', text).strip().title()
        return text if len(text) >= 2 and not any(c.isdigit() for c in text) else ""
'''
    def process_image(self, image_path):
        """Process single image with proper bounding box handling"""
        try:
            img = cv2.imread(image_path)
            if img is None:
                print(f"Could not read image: {image_path}")
                return {}, list(self.class_map.values())
            
            results = self.model(img)[0]
            extracted = {v: "" for v in self.class_map.values()}
            candidates = {k: [] for k in self.class_map.values()}

            for box, cls in zip(results.obb.xyxyxyxy, results.obb.cls):
                # Convert OBB to axis-aligned bounding box
                corners = box.cpu().numpy().reshape(4, 2).astype(np.int32)
                class_name = self.class_map[int(cls)]
                
                # Calculate bounding rectangle
                x, y, w, h = cv2.boundingRect(corners)
                x = max(0, min(x, img.shape[1]-1))
                y = max(0, min(y, img.shape[0]-1))
                w = max(1, min(w, img.shape[1]-x))
                h = max(1, min(h, img.shape[0]-y))
                
                # Crop and process region
                cropped = img[y:y+h, x:x+w]
                processed = self.preprocess(cropped)
                if processed is None:
                    continue
                
                text = pytesseract.image_to_string(processed, config='--psm 6 --oem 3').strip()
                print(text)  
                
                # Process text based on field type
                if class_name == "dob":
                    if dob := self._process_dob(text):
                        candidates["dob"].append(dob)
                elif class_name == "pan_number":
                    if pan := self._process_pan(text):
                        candidates["pan_number"].append(pan)
                else:
                    if processed_text := self._process_name(text):
                        candidates[class_name].append(processed_text)

            # Select best candidates
            for field in self.class_map.values():
                if candidates[field]:
                    if field == "pan_number":
                        # Select most common valid PAN format
                        valid_pans = [p for p in candidates[field] if len(p) == 10]
                        if valid_pans:
                            extracted[field] = max(set(valid_pans), key=valid_pans.count)
                    else:
                        # Select longest valid entry for other fields
                        extracted[field] = max(candidates[field], key=len, default="")

            missing = [field for field, value in extracted.items() if not value]
            return extracted, missing

        except Exception as e:
            print(f"Error processing {image_path}: {str(e)}")
            return {}, list(self.class_map.values())
        
    def process_batch(self, image_paths, output_file="pan_records.xlsx"):
        """Process multiple images and save to Excel"""
        try:
            wb, sheet = self._init_excel(output_file)
            success_count = 0
            
            for idx, img_path in enumerate(image_paths, 1):
                print(f"\nProcessing image {idx}/{len(image_paths)}: {os.path.basename(img_path)}")
                data, missing = self.process_image(img_path)
                
                if data:
                    '''print("\nExtracted Information:")
                    for k, v in data.items():
                        print(f"{k.upper():<12}: {v if v else 'Not Detected'}")'''

                    if missing:
                        print(f"\nMissing fields: {', '.join(missing)}")
                        if len(missing) >= 2:
                            print("❌Error: Try again by using clear image")
                        continue

                    self._write_excel_row(sheet, data)
                    success_count += 1
                    print("\nAll fields detected! Record saved to Excel.")

            self._finalize_excel(wb, output_file)
            print(f"\n✅Processing complete. Successful records: {success_count}/{len(image_paths)}")
            return True
        except Exception as e:
            print(f"❌Batch processing failed: {str(e)}")
            return False

    def _init_excel(self, output_file):
        """Initialize Excel workbook and sheet"""
        if os.path.exists(output_file):
            wb = load_workbook(output_file)
            sheet_name = "PAN Data"
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                if sheet.max_row == 1:
                    sheet.delete_rows(1)
            else:
                sheet = wb.create_sheet(sheet_name)
        else:
            wb = Workbook()
            sheet = wb.active
            sheet.title = "PAN Data"
            sheet.append(self.excel_headers)
        return wb, sheet

    def _write_excel_row(self, sheet, data):
        """Write single data row to Excel sheet"""
        row = [
            data.get("name", ""),
            data.get("father_name", ""),
            data.get("pan_number", ""),
            data.get("dob", "")
        ]
        sheet.append(row)

    def _finalize_excel(self, wb, output_file):
        """Auto-size columns and save workbook"""
        sheet = wb["PAN Data"]
        for col in range(1, len(self.excel_headers)+1):
            max_len = 0
            for row in range(1, sheet.max_row+1):
                cell_value = str(sheet.cell(row=row, column=col).value)
                max_len = max(max_len, len(cell_value))
            sheet.column_dimensions[get_column_letter(col)].width = max_len + 2
        wb.save(output_file)

if __name__ == "__main__":
    processor = PANProcessor()
    input_path = input("Enter image path or directory: ").strip('"')
    
    if os.path.isdir(input_path):
        image_paths = glob.glob(os.path.join(input_path, "*.[pj][np][gG]*")) + \
                     glob.glob(os.path.join(input_path, "*.[jJ][pP][eE][gG]*"))
    else:
        image_paths = [input_path]
    
    processor.process_batch(image_paths)
